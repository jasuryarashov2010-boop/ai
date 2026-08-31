import csv, io
from pathlib import Path
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

def extract_text(filename, data):
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        r = PdfReader(io.BytesIO(data))
        return "\n".join((p.extract_text() or "") for p in r.pages[:30]).strip()
    if ext == ".docx":
        d = Document(io.BytesIO(data))
        return "\n".join(p.text for p in d.paragraphs).strip()
    if ext in {".txt",".md",".py",".json",".xml",".html"}:
        return data.decode("utf-8", errors="replace")
    if ext == ".csv":
        rows = list(csv.reader(io.StringIO(data.decode("utf-8-sig", errors="replace"))))
        return "\n".join(" | ".join(r) for r in rows[:500])
    if ext == ".xlsx":
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        out = []
        for ws in wb.worksheets[:5]:
            out.append(f"[SHEET: {ws.title}]")
            for row in ws.iter_rows(max_row=500, values_only=True):
                vals = ["" if x is None else str(x) for x in row]
                if any(vals): out.append(" | ".join(vals))
        return "\n".join(out)
    raise ValueError("PDF, DOCX, XLSX, CSV, TXT, MD, PY yoki JSON yuboring")

def trim(text, limit=50000):
    return text if len(text) <= limit else text[:limit] + "\n\n[Matn qisqartirildi.]"
